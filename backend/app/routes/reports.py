"""
backend/app/routes/reports.py
Génération de rapports Excel avec openpyxl
pip install openpyxl --break-system-packages
"""
from flask import Blueprint, Response
from app.utils import role_required
from app.database import Database
from datetime import datetime, timedelta
import io

reports_bp = Blueprint('reports', __name__, url_prefix='/api/reports')


def get_stats(report_type: str) -> dict:
    now = datetime.now()

    if report_type == 'daily':
        date_filter = "DATE(entry_time) = CURDATE()"
        period = f"Journée du {now.strftime('%d/%m/%Y')}"
    elif report_type == 'weekly':
        date_filter = "entry_time >= DATE_SUB(NOW(), INTERVAL 7 DAY)"
        period = f"Du {(now - timedelta(days=7)).strftime('%d/%m/%Y')} au {now.strftime('%d/%m/%Y')}"
    elif report_type == 'monthly':
        date_filter = "MONTH(entry_time) = MONTH(NOW()) AND YEAR(entry_time) = YEAR(NOW())"
        period = f"Mois de {now.strftime('%m/%Y')}"
    else:
        date_filter = "entry_time >= DATE_SUB(NOW(), INTERVAL 30 DAY)"
        period = f"30 derniers jours au {now.strftime('%d/%m/%Y')}"

    def q(sql, fetch=False):
        return Database.execute_query(sql, fetch=fetch) if fetch else Database.execute_query_one(sql)

    entries = q(f"SELECT COUNT(*) AS cnt FROM parking_entries WHERE {date_filter}")
    exits   = q(f"SELECT COUNT(*) AS cnt FROM parking_entries WHERE status='OUT' AND {date_filter}")
    active  = q("SELECT COUNT(*) AS cnt FROM parking_entries WHERE status='IN'")
    users   = q("SELECT COUNT(*) AS cnt FROM users WHERE is_active=1")
    vehs    = q("SELECT COUNT(*) AS cnt FROM vehicles WHERE is_active=1")
    subs    = q("SELECT COUNT(*) AS cnt FROM subscriptions WHERE status='ACTIVE'")
    recs    = q("SELECT COUNT(*) AS cnt FROM reclamations WHERE status IN ('OPEN','IN_PROGRESS')")

    recent = Database.execute_query(
        f"""SELECT license_plate, vehicle_type, spot_number, entry_time, exit_time, status
            FROM parking_entries WHERE {date_filter}
            ORDER BY entry_time DESC LIMIT 50""",
        fetch=True
    ) or []

    all_users = Database.execute_query(
        "SELECT name, email, role, created_at FROM users WHERE is_active=1 ORDER BY created_at DESC",
        fetch=True
    ) or []

    all_subs = Database.execute_query(
        """SELECT s.license_plate, s.plan_type, s.start_date, s.end_date, s.status, u.name AS user_name
           FROM subscriptions s JOIN users u ON s.user_id = u.id
           ORDER BY s.created_at DESC LIMIT 50""",
        fetch=True
    ) or []

    return {
        'period': period, 'generated_at': now.strftime('%d/%m/%Y %H:%M'),
        'total_entries': entries['cnt'] if entries else 0,
        'total_exits':   exits['cnt']   if exits   else 0,
        'active_now':    active['cnt']  if active  else 0,
        'total_users':   users['cnt']   if users   else 0,
        'total_vehicles':vehs['cnt']    if vehs    else 0,
        'active_subs':   subs['cnt']    if subs    else 0,
        'open_recs':     recs['cnt']    if recs    else 0,
        'recent_entries': recent,
        'all_users': all_users,
        'all_subs':  all_subs,
    }


def generate_excel(stats: dict, title: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Couleurs ───────────────────────────────────────────────
    RED    = 'FFE53E3E'
    DARK   = 'FF1A1A1A'
    GRAY   = 'FF6B7280'
    LIGHT  = 'FFF9FAFB'
    WHITE  = 'FFFFFFFF'
    GREEN  = 'FF10B981'
    BLUE   = 'FF3B82F6'
    AMBER  = 'FFF59E0B'

    thin = Side(style='thin', color='FFE5E7EB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr_fill(color): return PatternFill('solid', fgColor=color)
    def font(bold=False, color='FF111827', size=10): return Font(bold=bold, color=color, size=size, name='Calibri')
    def center(): return Alignment(horizontal='center', vertical='center', wrap_text=True)
    def left():   return Alignment(horizontal='left',   vertical='center', wrap_text=True)

    # ══════════════════════════════════════════════════════════
    # FEUILLE 1 — Résumé
    # ══════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = 'Résumé'
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions['A'].width = 3
    ws1.column_dimensions['B'].width = 28
    ws1.column_dimensions['C'].width = 18
    ws1.column_dimensions['D'].width = 28
    ws1.column_dimensions['E'].width = 18
    ws1.column_dimensions['F'].width = 3
    ws1.row_dimensions[1].height = 8

    # Bannière header
    for row in range(2, 6):
        ws1.row_dimensions[row].height = 20 if row != 3 else 28
    for col in range(2, 6):
        cell = ws1.cell(row=2, column=col)
        cell.fill = hdr_fill(DARK)
    for col in range(2, 6):
        ws1.cell(row=3, column=col).fill = hdr_fill(DARK)
    for col in range(2, 6):
        ws1.cell(row=4, column=col).fill = hdr_fill(DARK)
    for col in range(2, 6):
        ws1.cell(row=5, column=col).fill = hdr_fill(DARK)

    ws1.merge_cells('B2:E5')
    c = ws1['B2']
    c.value = f'IMW PARKING — {title.upper()}'
    c.font = Font(bold=True, color=WHITE, size=18, name='Calibri')
    c.alignment = center()
    c.fill = hdr_fill(DARK)

    ws1.row_dimensions[6].height = 6

    # Période & généré
    ws1.merge_cells('B7:C7')
    ws1['B7'].value = f'Période : {stats["period"]}'
    ws1['B7'].font = font(bold=True, size=10, color='FF374151')
    ws1['B7'].alignment = left()

    ws1.merge_cells('D7:E7')
    ws1['D7'].value = f'Généré le : {stats["generated_at"]}'
    ws1['D7'].font = font(size=9, color=GRAY)
    ws1['D7'].alignment = Alignment(horizontal='right', vertical='center')

    ws1.row_dimensions[8].height = 8

    # KPI cards — ligne de titre
    ws1.row_dimensions[9].height = 22
    ws1.merge_cells('B9:E9')
    ws1['B9'].value = 'INDICATEURS CLÉS'
    ws1['B9'].font = Font(bold=True, color=WHITE, size=11, name='Calibri')
    ws1['B9'].fill = hdr_fill(RED)
    ws1['B9'].alignment = center()

    kpis = [
        ('Entrées enregistrées', stats['total_entries'], GREEN,  'Sorties enregistrées', stats['total_exits'],    BLUE),
        ('Véhicules actifs',     stats['active_now'],    AMBER,  'Utilisateurs inscrits', stats['total_users'],   GRAY),
        ('Abonnements actifs',   stats['active_subs'],   GREEN,  'Réclamations ouvertes', stats['open_recs'],     RED),
        ('Véhicules enregistrés',stats['total_vehicles'],BLUE,   'Taux d\'occupation',   f'{round((stats["active_now"]/160)*100)}%', AMBER),
    ]

    row = 10
    for label1, val1, c1, label2, val2, c2 in kpis:
        ws1.row_dimensions[row].height = 14
        ws1.row_dimensions[row+1].height = 22

        # Carte gauche
        ws1.merge_cells(f'B{row}:C{row}')
        ws1[f'B{row}'].value = label1
        ws1[f'B{row}'].font = font(size=8, color=GRAY)
        ws1[f'B{row}'].alignment = left()
        ws1[f'B{row}'].fill = hdr_fill(LIGHT)
        ws1[f'C{row}'].fill = hdr_fill(LIGHT)

        ws1.merge_cells(f'B{row+1}:C{row+1}')
        ws1[f'B{row+1}'].value = val1
        ws1[f'B{row+1}'].font = Font(bold=True, size=16, color=c1, name='Calibri')
        ws1[f'B{row+1}'].alignment = left()
        ws1[f'B{row+1}'].fill = hdr_fill(LIGHT)
        ws1[f'C{row+1}'].fill = hdr_fill(LIGHT)

        # Carte droite
        ws1.merge_cells(f'D{row}:E{row}')
        ws1[f'D{row}'].value = label2
        ws1[f'D{row}'].font = font(size=8, color=GRAY)
        ws1[f'D{row}'].alignment = left()
        ws1[f'D{row}'].fill = hdr_fill(LIGHT)
        ws1[f'E{row}'].fill = hdr_fill(LIGHT)

        ws1.merge_cells(f'D{row+1}:E{row+1}')
        ws1[f'D{row+1}'].value = val2
        ws1[f'D{row+1}'].font = Font(bold=True, size=16, color=c2, name='Calibri')
        ws1[f'D{row+1}'].alignment = left()
        ws1[f'D{row+1}'].fill = hdr_fill(LIGHT)
        ws1[f'E{row+1}'].fill = hdr_fill(LIGHT)

        row += 3

    # ══════════════════════════════════════════════════════════
    # FEUILLE 2 — Entrées / Sorties
    # ══════════════════════════════════════════════════════════
    ws2 = wb.create_sheet('Entrées & Sorties')
    ws2.sheet_view.showGridLines = False

    cols2 = ['Plaque', 'Type Véhicule', 'Place', "Heure d'entrée", 'Heure de sortie', 'Statut']
    widths2 = [18, 16, 10, 22, 22, 12]
    for i, w in enumerate(widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # Header bannière
    ws2.merge_cells(f'A1:{get_column_letter(len(cols2))}1')
    ws2['A1'].value = f'ENTRÉES & SORTIES — {stats["period"]}'
    ws2['A1'].font = Font(bold=True, color=WHITE, size=13, name='Calibri')
    ws2['A1'].fill = hdr_fill(DARK)
    ws2['A1'].alignment = center()
    ws2.row_dimensions[1].height = 28

    # Colonnes
    ws2.row_dimensions[2].height = 20
    for i, col in enumerate(cols2, 1):
        c = ws2.cell(row=2, column=i, value=col)
        c.font = Font(bold=True, color=WHITE, size=9, name='Calibri')
        c.fill = hdr_fill(RED)
        c.alignment = center()
        c.border = border

    for ridx, entry in enumerate(stats['recent_entries'], 3):
        ws2.row_dimensions[ridx].height = 18
        fill = hdr_fill(WHITE) if ridx % 2 == 1 else hdr_fill(LIGHT)
        row_data = [
            str(entry.get('license_plate', '—')),
            str(entry.get('vehicle_type', '—')),
            str(entry.get('spot_number', '—')) if entry.get('spot_number') else '—',
            str(entry.get('entry_time', '—'))[:16] if entry.get('entry_time') else '—',
            str(entry.get('exit_time', '—'))[:16]  if entry.get('exit_time')  else '—',
            'Présent' if entry.get('status') == 'IN' else 'Sorti',
        ]
        for cidx, val in enumerate(row_data, 1):
            c = ws2.cell(row=ridx, column=cidx, value=val)
            c.font = font(size=9)
            c.alignment = center()
            c.fill = fill
            c.border = border
            if cidx == 6:
                c.font = Font(bold=True, size=9, name='Calibri',
                              color=GREEN if val == 'Présent' else GRAY)

    # ══════════════════════════════════════════════════════════
    # FEUILLE 3 — Utilisateurs
    # ══════════════════════════════════════════════════════════
    ws3 = wb.create_sheet('Utilisateurs')
    ws3.sheet_view.showGridLines = False

    cols3 = ['Nom', 'Email', 'Rôle', "Date d'inscription"]
    widths3 = [24, 32, 14, 22]
    for i, w in enumerate(widths3, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    ws3.merge_cells(f'A1:{get_column_letter(len(cols3))}1')
    ws3['A1'].value = f'UTILISATEURS INSCRITS — {stats["generated_at"]}'
    ws3['A1'].font = Font(bold=True, color=WHITE, size=13, name='Calibri')
    ws3['A1'].fill = hdr_fill(DARK)
    ws3['A1'].alignment = center()
    ws3.row_dimensions[1].height = 28

    ws3.row_dimensions[2].height = 20
    role_colors = {'ADMIN': RED, 'MANAGER': BLUE, 'AGENT': AMBER, 'CLIENT': GREEN}

    for i, col in enumerate(cols3, 1):
        c = ws3.cell(row=2, column=i, value=col)
        c.font = Font(bold=True, color=WHITE, size=9, name='Calibri')
        c.fill = hdr_fill(RED)
        c.alignment = center()
        c.border = border

    for ridx, user in enumerate(stats['all_users'], 3):
        ws3.row_dimensions[ridx].height = 18
        fill = hdr_fill(WHITE) if ridx % 2 == 1 else hdr_fill(LIGHT)
        row_data = [
            str(user.get('name', '—')),
            str(user.get('email', '—')),
            str(user.get('role', '—')),
            str(user.get('created_at', '—'))[:10] if user.get('created_at') else '—',
        ]
        for cidx, val in enumerate(row_data, 1):
            c = ws3.cell(row=ridx, column=cidx, value=val)
            c.font = font(size=9)
            c.alignment = center()
            c.fill = fill
            c.border = border
            if cidx == 3:
                role_color = role_colors.get(val, GRAY)
                c.font = Font(bold=True, size=9, color=role_color, name='Calibri')

    # ══════════════════════════════════════════════════════════
    # FEUILLE 4 — Abonnements
    # ══════════════════════════════════════════════════════════
    ws4 = wb.create_sheet('Abonnements')
    ws4.sheet_view.showGridLines = False

    cols4 = ['Client', 'Plaque', 'Plan', 'Début', 'Fin', 'Statut']
    widths4 = [24, 16, 12, 14, 14, 14]
    for i, w in enumerate(widths4, 1):
        ws4.column_dimensions[get_column_letter(i)].width = w

    ws4.merge_cells(f'A1:{get_column_letter(len(cols4))}1')
    ws4['A1'].value = f'ABONNEMENTS — {stats["generated_at"]}'
    ws4['A1'].font = Font(bold=True, color=WHITE, size=13, name='Calibri')
    ws4['A1'].fill = hdr_fill(DARK)
    ws4['A1'].alignment = center()
    ws4.row_dimensions[1].height = 28

    ws4.row_dimensions[2].height = 20
    status_colors = {'ACTIVE': GREEN, 'EXPIRED': AMBER, 'CANCELLED': RED}

    for i, col in enumerate(cols4, 1):
        c = ws4.cell(row=2, column=i, value=col)
        c.font = Font(bold=True, color=WHITE, size=9, name='Calibri')
        c.fill = hdr_fill(RED)
        c.alignment = center()
        c.border = border

    for ridx, sub in enumerate(stats['all_subs'], 3):
        ws4.row_dimensions[ridx].height = 18
        fill = hdr_fill(WHITE) if ridx % 2 == 1 else hdr_fill(LIGHT)
        status = str(sub.get('status', '—'))
        row_data = [
            str(sub.get('user_name', '—')),
            str(sub.get('license_plate', '—')),
            str(sub.get('plan_type', '—')),
            str(sub.get('start_date', '—'))[:10] if sub.get('start_date') else '—',
            str(sub.get('end_date', '—'))[:10]   if sub.get('end_date')   else '—',
            status,
        ]
        for cidx, val in enumerate(row_data, 1):
            c = ws4.cell(row=ridx, column=cidx, value=val)
            c.font = font(size=9)
            c.alignment = center()
            c.fill = fill
            c.border = border
            if cidx == 6:
                c.font = Font(bold=True, size=9, color=status_colors.get(val, GRAY), name='Calibri')

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


@reports_bp.route('/<string:report_type>', methods=['GET'])
@role_required(['ADMIN', 'MANAGER'])
def download_report(report_type: str):
    titles = {
        'daily':    'Rapport Quotidien',
        'weekly':   'Rapport Hebdomadaire',
        'monthly':  'Analyse Mensuelle',
        'security': 'Audit Sécurité',
    }
    if report_type not in titles:
        return {'error': 'Type invalide'}, 400

    stats = get_stats(report_type)
    xlsx  = generate_excel(stats, titles[report_type])
    fname = f"IMW_Parking_{report_type}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    return Response(
        xlsx,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{fname}"'}
    )